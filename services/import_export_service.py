"""导入导出服务"""
from __future__ import annotations

from datetime import datetime, date
from typing import Optional

from services.todo_service import TodoService
from services.category_service import CategoryService
from models.todo import Todo
from config.constants import STATUS_TODO, STATUS_DONE, STATUS_ARCHIVED

STATUS_LABELS = {STATUS_TODO: "未完成", STATUS_DONE: "已完成", STATUS_ARCHIVED: "已归档"}


EXPORT_VERSION = "2.0"

EXPORT_FIELDS_TODO = {
    "title", "description", "priority", "status", "color_tag",
    "start_date", "due_date", "auto_postpone", "sort_order", "category_name",
    "recurrence_type", "recurrence_interval", "recurrence_day",
    "recurrence_start_date", "recurrence_end_date", "is_recurrence_template",
    "occurrence_date", "is_exception", "completed_at",
}

EXPORT_FIELDS_CHILD = {
    "title", "description", "priority", "status", "color_tag",
    "sort_order", "category_name", "completed_at",
}

EXPORT_FIELDS_INSTANCE = {
    "title", "description", "priority", "status", "color_tag",
    "start_date", "due_date", "sort_order", "category_name",
    "occurrence_date", "is_exception", "completed_at",
}


class ImportExportService:

    def __init__(self, todo_service: TodoService, category_service: CategoryService):
        self.todo_service = todo_service
        self.category_service = category_service

    # ---- 导出 ----

    def export_data(self) -> dict:
        categories = self.category_service.get_all()
        all_todos = self.todo_service.get_all_including_archived()

        cat_list = []
        for cat in categories:
            if not cat.is_system:
                cat_list.append({
                    "name": cat.name,
                    "color": cat.color,
                    "sort_order": cat.sort_order,
                })

        todo_list = self._build_export_tree(all_todos)

        return {
            "version": EXPORT_VERSION,
            "app": "EasyTodo",
            "exported_at": datetime.now().isoformat(),
            "categories": cat_list,
            "todos": todo_list,
        }

    def export_to_excel(
        self,
        path: str,
        date_field: Optional[str] = None,
        start_date: Optional[date] = None,
        end_date: Optional[date] = None,
    ) -> int:
        """导出数据到 Excel 文件，返回导出的任务数量

        :param path: 导出文件路径
        :param date_field: 时间筛选字段，支持
            start_date / due_date / completed_at / created_at；
            为 None 或空字符串表示不按时间筛选，导出全部
        :param start_date: 起始日期（含），None 表示不限
        :param end_date: 结束日期（含），None 表示不限
        """
        from openpyxl import Workbook
        from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
        from openpyxl.utils import get_column_letter

        all_todos = self.todo_service.get_all_including_archived()

        children_map: dict[int, list[Todo]] = {}
        for t in all_todos:
            if t.pid is not None:
                children_map.setdefault(t.pid, []).append(t)

        def _matches_filter(t: Todo) -> bool:
            if not date_field:
                return True
            value = getattr(t, date_field, None)
            if value is None:
                return False
            if isinstance(value, datetime):
                cmp_value = value.date()
            else:
                cmp_value = value
            if start_date and cmp_value < start_date:
                return False
            if end_date and cmp_value > end_date:
                return False
            return True

        wb = Workbook()
        ws = wb.active
        ws.title = "任务数据"

        headers = ["任务名称", "详情", "开始时间", "完成时间", "创建时间", "更新时间", "分类", "状态", "任务类型"]
        ws.append(headers)

        header_fill = PatternFill(start_color="0078D4", end_color="0078D4", fill_type="solid")
        header_font_white = Font(bold=True, size=11, color="FFFFFF")
        thin_border = Border(
            left=Side(style="thin", color="D0D0D0"),
            right=Side(style="thin", color="D0D0D0"),
            top=Side(style="thin", color="D0D0D0"),
            bottom=Side(style="thin", color="D0D0D0"),
        )
        wrap_alignment = Alignment(wrap_text=True, vertical="top")

        for col_idx, _ in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col_idx)
            cell.font = header_font_white
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal="center", vertical="center")
            cell.border = thin_border

        row_idx = 2
        exported_count = 0

        for t in all_todos:
            if t.pid is not None:
                continue
            if not _matches_filter(t):
                continue
            is_recurring = bool(t.recurrence_type)
            task_type = "重复任务" if is_recurring else "普通任务"
            status_label = STATUS_LABELS.get(t.status, "未完成")
            cat_name = t.category.name if t.category else ""

            row_data = [
                t.title or "",
                t.description or "",
                t.start_date.isoformat() if t.start_date else "",
                t.completed_at.strftime("%Y-%m-%d %H:%M:%S") if t.completed_at else "",
                t.created_at.strftime("%Y-%m-%d %H:%M:%S") if t.created_at else "",
                t.updated_at.strftime("%Y-%m-%d %H:%M:%S") if t.updated_at else "",
                cat_name,
                status_label,
                task_type,
            ]
            ws.append(row_data)
            for col_idx in range(1, len(headers) + 1):
                cell = ws.cell(row=row_idx, column=col_idx)
                cell.border = thin_border
                cell.alignment = wrap_alignment
            row_idx += 1
            exported_count += 1

            children = sorted(children_map.get(t.id, []), key=lambda x: x.sort_order)
            for child in children:
                child_row = [child.title or "", "", "", "", "", "", "", "", ""]
                ws.append(child_row)
                for col_idx in range(1, len(headers) + 1):
                    cell = ws.cell(row=row_idx, column=col_idx)
                    cell.border = thin_border
                    cell.alignment = wrap_alignment
                    cell.font = Font(color="666666")
                row_idx += 1

        col_widths = [30, 40, 16, 20, 20, 20, 12, 10, 12]
        for i, width in enumerate(col_widths, 1):
            ws.column_dimensions[get_column_letter(i)].width = width

        wb.save(path)
        return exported_count

    def _build_export_tree(self, todos: list[Todo]) -> list[dict]:
        id_to_todo = {t.id: t for t in todos}

        children_map: dict[int, list[Todo]] = {}
        instances_map: dict[int, list[Todo]] = {}
        instance_children_map: dict[int, list[Todo]] = {}

        for t in todos:
            if t.pid is not None:
                parent = id_to_todo.get(t.pid)
                if parent is None:
                    continue
                if parent.is_recurrence_template and parent.pid is None:
                    pass
                elif parent.recurrence_template_id is not None:
                    instance_children_map.setdefault(parent.id, []).append(t)
                else:
                    children_map.setdefault(parent.id, []).append(t)

        for t in todos:
            if t.recurrence_template_id is not None and t.pid is None:
                instances_map.setdefault(t.recurrence_template_id, []).append(t)

        result = []
        for t in todos:
            if t.pid is not None:
                continue
            if t.is_recurrence_template:
                node = self._todo_to_export_node(
                    t, EXPORT_FIELDS_TODO,
                    children_map.get(t.id, []),
                    instances_map.get(t.id, []),
                    instance_children_map,
                )
            else:
                node = self._todo_to_export_node(
                    t, EXPORT_FIELDS_TODO,
                    children_map.get(t.id, []),
                    None,
                    instance_children_map,
                )
            result.append(node)

        return result

    def _todo_to_export_node(
        self,
        todo: Todo,
        fields: set[str],
        children: list[Todo],
        instances: Optional[list[Todo]],
        instance_children_map: dict[int, list[Todo]],
    ) -> dict:
        base = todo.to_export_dict()
        node = {k: v for k, v in base.items() if k in fields}

        node["children"] = [
            self._child_to_export_node(c, instance_children_map)
            for c in sorted(children, key=lambda x: x.sort_order)
        ]

        if instances is not None:
            node["instances"] = [
                self._instance_to_export_node(inst, instance_children_map.get(inst.id, []))
                for inst in sorted(instances, key=lambda x: x.occurrence_date or date.min)
            ]

        return node

    def _child_to_export_node(
        self,
        child: Todo,
        instance_children_map: dict[int, list[Todo]],
    ) -> dict:
        base = child.to_export_dict()
        node = {k: v for k, v in base.items() if k in EXPORT_FIELDS_CHILD}

        sub_children = instance_children_map.get(child.id, [])

        node["children"] = [
            self._child_to_export_node(c, instance_children_map)
            for c in sorted(sub_children, key=lambda x: x.sort_order)
        ]
        return node

    def _instance_to_export_node(
        self,
        instance: Todo,
        instance_children: list[Todo],
    ) -> dict:
        base = instance.to_export_dict()
        node = {k: v for k, v in base.items() if k in EXPORT_FIELDS_INSTANCE}

        node["children"] = [
            self._child_to_export_node(c, {})
            for c in sorted(instance_children, key=lambda x: x.sort_order)
        ]
        return node

    # ---- 导入预览 ----

    def preview(self, data) -> dict:
        if isinstance(data, list):
            return self._preview_v1(data)
        if isinstance(data, dict) and data.get("version") == EXPORT_VERSION:
            return self._preview_v2(data)
        if isinstance(data, dict) and "todos" in data:
            return self._preview_v2(data)
        return {"valid": False, "error": "文件格式不正确"}

    def _preview_v1(self, data: list) -> dict:
        if not data:
            return {"valid": False, "error": "文件中没有数据"}
        cat_names = set()
        todo_count = 0
        child_count = 0
        for item in data:
            title = item.get("title", "").strip()
            if not title:
                continue
            todo_count += 1
            cat_info = item.get("category")
            if isinstance(cat_info, dict):
                name = cat_info.get("name", "")
                if name:
                    cat_names.add(name)
            cat_name = item.get("category_name")
            if cat_name:
                cat_names.add(cat_name)
            children = item.get("children", [])
            child_count += len(children)
        return {
            "valid": True,
            "version": "1.0",
            "categories_count": len(cat_names),
            "todos_count": todo_count,
            "children_count": child_count,
            "instances_count": 0,
            "duplicate_count": 0,
        }

    def _preview_v2(self, data: dict) -> dict:
        categories = data.get("categories", [])
        todos = data.get("todos", [])
        todo_count = 0
        child_count = 0
        instance_count = 0
        for t in todos:
            todo_count += 1
            child_count += self._count_children(t.get("children", []))
            instances = t.get("instances", [])
            instance_count += len(instances)
            for inst in instances:
                child_count += self._count_children(inst.get("children", []))
        return {
            "valid": True,
            "version": data.get("version", "2.0"),
            "categories_count": len(categories),
            "todos_count": todo_count,
            "children_count": child_count,
            "instances_count": instance_count,
            "duplicate_count": 0,
        }

    def _count_children(self, children: list) -> int:
        count = len(children)
        for c in children:
            count += self._count_children(c.get("children", []))
        return count

    # ---- 导入 ----

    def import_data(self, data, mode: str = "append") -> dict:
        if isinstance(data, list):
            return self._import_v1(data, mode)
        if isinstance(data, dict):
            version = data.get("version", "")
            if version == EXPORT_VERSION or "todos" in data:
                return self._import_v2(data, mode)
        raise ValueError("不支持的文件格式")

    def _import_v1(self, data: list, mode: str) -> dict:
        if mode == "replace":
            self._clear_all_data()

        cat_name_map = self._build_category_name_map()

        count = 0
        for item in data:
            title = item.get("title", "").strip()
            if not title:
                continue

            cat_info = item.pop("category", None)
            if cat_info and isinstance(cat_info, dict):
                cat_name = cat_info.get("name", "")
                if cat_name:
                    item["category_id"] = self._match_or_create_category(cat_name, cat_name_map)

            cat_name = item.get("category_name")
            if cat_name:
                item["category_id"] = self._match_or_create_category(cat_name, cat_name_map)
                item.pop("category_name", None)

            due = item.get("due_date")
            if isinstance(due, str) and due:
                try:
                    item["due_date"] = date.fromisoformat(due)
                except Exception:
                    item["due_date"] = None
            elif not isinstance(due, date):
                item["due_date"] = None

            for key in ("id", "pid", "created_at", "updated_at", "sort_order",
                        "children", "is_recurrence_template",
                        "recurrence_template_id", "occurrence_date", "is_exception",
                        "recurrence_type", "recurrence_interval",
                        "recurrence_day", "recurrence_end_date", "status",
                        "completed_at"):
                item.pop(key, None)

            self.todo_service.create(**item)
            count += 1

        return {"imported": count, "categories": 0}

    def _import_v2(self, data: dict, mode: str) -> dict:
        if mode == "replace":
            self._clear_all_data()

        cat_name_map = self._build_category_name_map()

        imported_cats = self._import_categories(data.get("categories", []), cat_name_map)

        count = 0
        children_count = 0
        instance_count = 0

        for todo_node in data.get("todos", []):
            is_template = todo_node.get("is_recurrence_template", False)
            if is_template:
                created, ch, inst = self._import_template_node(todo_node, cat_name_map)
            else:
                created, ch = self._import_todo_node(todo_node, cat_name_map)
                inst = 0

            if created:
                count += 1
                children_count += ch
                instance_count += inst

        return {
            "imported": count,
            "children": children_count,
            "instances": instance_count,
            "categories": imported_cats,
        }

    def _import_categories(self, categories: list[dict], cat_name_map: dict[str, int]) -> int:
        count = 0
        for cat_data in categories:
            name = cat_data.get("name", "").strip()
            if not name:
                continue
            if name not in cat_name_map:
                color = cat_data.get("color", "#0078D4")
                new_cat = self.category_service.create(name, color)
                cat_name_map[name] = new_cat.id
                count += 1
        return count

    def _import_todo_node(self, node: dict, cat_name_map: dict[str, int], pid: int = None) -> tuple[bool, int]:
        params = self._node_to_raw_params(node, cat_name_map, pid=pid)

        todo = self.todo_service.create_raw(**params)
        children_count = self._import_children(node.get("children", []), cat_name_map, todo.id)
        return True, children_count

    def _import_template_node(self, node: dict, cat_name_map: dict[str, int]) -> tuple[bool, int, int]:
        params = self._node_to_raw_params(node, cat_name_map)
        params["is_recurrence_template"] = True

        template = self.todo_service.create_raw(**params)

        blueprint_children = node.get("children", [])
        for child_node in blueprint_children:
            child_params = self._node_to_raw_params(child_node, cat_name_map, pid=template.id)
            self.todo_service.create_raw(**child_params)

        instance_count = 0
        for inst_node in node.get("instances", []):
            inst_params = self._instance_to_raw_params(inst_node, cat_name_map, template.id)
            instance = self.todo_service.create_raw(**inst_params)
            for inst_child in inst_node.get("children", []):
                ic_params = self._node_to_raw_params(inst_child, cat_name_map, pid=instance.id)
                self.todo_service.create_raw(**ic_params)
            instance_count += 1

        return True, len(blueprint_children), instance_count

    def _import_children(self, children: list[dict], cat_name_map: dict[str, int], parent_id: int) -> int:
        count = 0
        for child_node in children:
            params = self._node_to_raw_params(child_node, cat_name_map, pid=parent_id)
            child = self.todo_service.create_raw(**params)
            count += 1
            sub_children = child_node.get("children", [])
            if sub_children:
                count += self._import_children(sub_children, cat_name_map, child.id)
        return count

    def _node_to_raw_params(self, node: dict, cat_name_map: dict[str, int], pid: int = None) -> dict:
        params = {}

        simple_fields = [
            "title", "description", "priority", "status", "color_tag",
            "auto_postpone", "sort_order",
            "recurrence_type", "recurrence_interval",
            "recurrence_day",
        ]
        for field in simple_fields:
            if field in node:
                params[field] = node[field]

        completed = node.get("completed_at")
        if isinstance(completed, str) and completed:
            try:
                params["completed_at"] = datetime.fromisoformat(completed)
            except Exception:
                params["completed_at"] = None
        elif isinstance(completed, datetime):
            params["completed_at"] = completed
        else:
            params["completed_at"] = None

        if pid is not None:
            params["pid"] = pid

        cat_name = node.get("category_name")
        if cat_name:
            params["category_id"] = self._match_or_create_category(cat_name, cat_name_map)

        due = node.get("due_date")
        if isinstance(due, str) and due:
            try:
                params["due_date"] = date.fromisoformat(due)
            except Exception:
                params["due_date"] = None
        elif isinstance(due, date):
            params["due_date"] = due
        else:
            params["due_date"] = None

        start = node.get("start_date")
        if isinstance(start, str) and start:
            try:
                params["start_date"] = date.fromisoformat(start)
            except Exception:
                params["start_date"] = None
        elif isinstance(start, date):
            params["start_date"] = start
        else:
            params["start_date"] = None

        end_date = node.get("recurrence_end_date")
        if isinstance(end_date, str) and end_date:
            try:
                params["recurrence_end_date"] = date.fromisoformat(end_date)
            except Exception:
                params["recurrence_end_date"] = None
        elif isinstance(end_date, date):
            params["recurrence_end_date"] = end_date
        else:
            params["recurrence_end_date"] = None

        start_date = node.get("recurrence_start_date")
        if isinstance(start_date, str) and start_date:
            try:
                params["recurrence_start_date"] = date.fromisoformat(start_date)
            except Exception:
                params["recurrence_start_date"] = None
        elif isinstance(start_date, date):
            params["recurrence_start_date"] = start_date
        else:
            params["recurrence_start_date"] = None

        return params

    def _instance_to_raw_params(self, node: dict, cat_name_map: dict[str, int], template_id: int) -> dict:
        params = {
            "title": node.get("title", ""),
            "description": node.get("description", ""),
            "priority": node.get("priority", 0),
            "color_tag": node.get("color_tag"),
            "sort_order": node.get("sort_order", 0),
            "recurrence_template_id": template_id,
            "is_exception": node.get("is_exception", False),
            "is_recurrence_template": False,
        }

        completed = node.get("completed_at")
        if isinstance(completed, str) and completed:
            try:
                params["completed_at"] = datetime.fromisoformat(completed)
            except Exception:
                params["completed_at"] = None
        elif isinstance(completed, datetime):
            params["completed_at"] = completed
        else:
            params["completed_at"] = None

        status = node.get("status")
        if status is not None:
            params["status"] = status

        cat_name = node.get("category_name")
        if cat_name:
            params["category_id"] = self._match_or_create_category(cat_name, cat_name_map)

        due = node.get("due_date")
        if isinstance(due, str) and due:
            try:
                params["due_date"] = date.fromisoformat(due)
            except Exception:
                params["due_date"] = None
        elif isinstance(due, date):
            params["due_date"] = due
        else:
            params["due_date"] = None

        start = node.get("start_date")
        if isinstance(start, str) and start:
            try:
                params["start_date"] = date.fromisoformat(start)
            except Exception:
                params["start_date"] = None
        elif isinstance(start, date):
            params["start_date"] = start
        else:
            params["start_date"] = None

        occ = node.get("occurrence_date")
        if isinstance(occ, str) and occ:
            try:
                params["occurrence_date"] = date.fromisoformat(occ)
            except Exception:
                params["occurrence_date"] = None
        elif isinstance(occ, date):
            params["occurrence_date"] = occ
        else:
            params["occurrence_date"] = None

        return params

    def _build_category_name_map(self) -> dict[str, int]:
        result = {}
        for cat in self.category_service.get_all():
            result[cat.name] = cat.id
        return result

    def _match_or_create_category(self, name: str, cat_name_map: dict[str, int]) -> Optional[int]:
        if not name:
            return None
        if name in cat_name_map:
            return cat_name_map[name]
        new_cat = self.category_service.create(name)
        cat_name_map[name] = new_cat.id
        return new_cat.id

    def _clear_all_data(self):
        from models.database import db
        session = db.get_session()
        try:
            session.query(Todo).filter(Todo.pid.isnot(None)).delete(synchronize_session=False)
            session.query(Todo).filter(Todo.pid.is_(None)).delete(synchronize_session=False)
            from models.category import Category
            session.query(Category).filter(Category.is_system == False).delete(synchronize_session=False)
            session.commit()
        except Exception:
            session.rollback()
            raise
        finally:
            session.close()
